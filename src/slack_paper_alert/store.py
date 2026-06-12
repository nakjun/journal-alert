from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .abstract_fetchers import fetch_direct_abstract
from .config import DEFAULT_EXCEL_PATH, DEFAULT_MARKDOWN_PATH
from .models import Paper


HEADERS = [
    "discovered_date",
    "published_date",
    "journal",
    "modality",
    "method_family",
    "title",
    "authors",
    "doi",
    "url",
    "abstract",
    "source_api",
    "source_journal_config",
]


def append_papers(
    papers: list[Paper],
    excel_path: Path = DEFAULT_EXCEL_PATH,
    markdown_path: Path = DEFAULT_MARKDOWN_PATH,
) -> list[Paper]:
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb, ws = _load_or_create_workbook(excel_path)
    existing = _existing_keys(ws)

    added: list[Paper] = []
    for paper in papers:
        if paper.unique_key in existing:
            continue
        ws.append([getattr(paper, header) for header in HEADERS])
        existing.add(paper.unique_key)
        added.append(paper)

    _format_sheet(ws)
    wb.save(excel_path)
    write_markdown_from_excel(excel_path, markdown_path)
    return added


def write_markdown_from_excel(
    excel_path: Path = DEFAULT_EXCEL_PATH,
    markdown_path: Path = DEFAULT_MARKDOWN_PATH,
) -> None:
    if not excel_path.exists():
        return

    wb = load_workbook(excel_path)
    ws = wb["papers"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    visible_headers = [
        "discovered_date",
        "published_date",
        "journal",
        "modality",
        "method_family",
        "title",
        "authors",
        "url",
        "abstract",
    ]
    indexes = [HEADERS.index(header) for header in visible_headers]

    lines = [
        "# Paper Database",
        "",
        f"Total records: {max(len(rows) - 1, 0)}",
        "",
        "| " + " | ".join(visible_headers) + " |",
        "| " + " | ".join("---" for _ in visible_headers) + " |",
    ]
    for row in rows[1:]:
        values = [_md_cell(row[index]) for index in indexes]
        lines.append("| " + " | ".join(values) + " |")

    markdown_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def refresh_missing_abstracts(
    excel_path: Path = DEFAULT_EXCEL_PATH,
    markdown_path: Path = DEFAULT_MARKDOWN_PATH,
    force: bool = False,
) -> int:
    if not excel_path.exists():
        return 0

    wb = load_workbook(excel_path)
    ws = wb["papers"]
    header_row = [cell.value for cell in ws[1]]
    indexes = {header: header_row.index(header) + 1 for header in HEADERS if header in header_row}
    updated = 0

    for row in range(2, ws.max_row + 1):
        current = str(ws.cell(row=row, column=indexes["abstract"]).value or "").strip()
        if current and not force:
            continue

        result = fetch_direct_abstract(
            url=str(ws.cell(row=row, column=indexes["url"]).value or ""),
            doi=str(ws.cell(row=row, column=indexes["doi"]).value or ""),
            journal=str(ws.cell(row=row, column=indexes["journal"]).value or ""),
        )
        if not result:
            continue
        if current and len(result.abstract) <= len(current):
            continue

        ws.cell(row=row, column=indexes["abstract"]).value = result.abstract
        ws.cell(row=row, column=indexes["url"]).value = result.final_url
        updated += 1

    if updated:
        _format_sheet(ws)
        wb.save(excel_path)
    write_markdown_from_excel(excel_path, markdown_path)
    return updated


def _load_or_create_workbook(path: Path) -> tuple[Workbook, Worksheet]:
    if path.exists():
        wb = load_workbook(path)
        if "papers" in wb.sheetnames:
            ws = wb["papers"]
        else:
            ws = wb.active
            ws.title = "papers"
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            ws.append(HEADERS)
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "papers"
    ws.append(HEADERS)
    return wb, ws


def _existing_keys(ws: Worksheet) -> set[str]:
    keys: set[str] = set()
    header_row = [cell.value for cell in ws[1]]
    try:
        doi_idx = header_row.index("doi") + 1
        title_idx = header_row.index("title") + 1
        journal_idx = header_row.index("journal") + 1
    except ValueError:
        return keys

    for row in range(2, ws.max_row + 1):
        doi = str(ws.cell(row=row, column=doi_idx).value or "").strip().lower()
        title = str(ws.cell(row=row, column=title_idx).value or "").strip().lower()
        journal = str(ws.cell(row=row, column=journal_idx).value or "").strip().lower()
        keys.add(doi or f"{title}|{journal}")
    return keys


def _format_sheet(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    widths = {
        "A": 16,
        "B": 16,
        "C": 34,
        "D": 18,
        "E": 24,
        "F": 70,
        "G": 42,
        "H": 28,
        "I": 48,
        "J": 90,
        "K": 14,
        "L": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _md_cell(value: object) -> str:
    text = str(value or "").replace("\n", " ").replace("|", "\\|").strip()
    return text
